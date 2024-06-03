import os
import time
import subprocess

from openpyxl.styles import Alignment, Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import pyautogui
from bs4 import BeautifulSoup

# List of Pokémon to scrape data for. If empty, data for all Pokémon will be scraped.
pokemon_list = []  # Add Pokémon names here if needed
pokemon_list = [pokemon.lower() for pokemon in pokemon_list]  # Convert to lowercase for case-insensitive comparison

# Path to the file where the HTML content will be saved
html_file_path = 'html_content.html'


def check_install_dependencies():
    print("Checking and installing dependencies...")
    dependencies = ["selenium", "openpyxl", "beautifulsoup4", "lxml", "pyautogui"]
    for dependency in dependencies:
        try:
            __import__(dependency)
            print(f"{dependency} is installed.")
        except ImportError:
            print(f"{dependency} is not installed. Installing now...")
            subprocess.check_call(["python", "-m", "pip", "install", dependency])


def extract_pokemon_info(row):
    dex_id = int(row.find('td', class_='speciesDexIDWrapper').text)
    name = row.find('td', class_='speciesNameWrapper').text
    types = ', '.join([t.text for t in row.find_all('div', class_='typeWrapper')])
    primary_ability = row.find('div', class_='speciesAbilitiesPrimary').text
    if row.find('div', class_='speciesAbilitiesSecondary'):
        secondary_ability = row.find('div', class_='speciesAbilitiesSecondary').text
    else:
        secondary_ability = "-"
    if row.find('div', class_='speciesAbilitiesHidden'):
        hidden_ability = row.find('div', class_='speciesAbilitiesHidden').text
    else:
        hidden_ability = "-"
    stats = [int(stat.text) for stat in row.find_all('div', class_='speciesStatValue')]
    sprite_base64 = row.find('img', class_='speciesSprite')['src'].split(',')[1]  # Extract base64 code
    return [dex_id, name, types, primary_ability, secondary_ability, hidden_ability, *stats, sprite_base64]


def scrape_pokemon(html_content):
    print("Scraping Pokémon data...")
    # Parse the HTML content
    soup = BeautifulSoup(html_content, 'lxml')

    # Find all the rows containing Pokémon data
    pokemon_rows = soup.find_all('tr', class_='speciesRow')

    # Create a workbook
    wb = Workbook()
    ws = wb.active

    # Define header
    header = ["DexID", "Name", "Types", "Primary Ability", "Secondary Ability", "Hidden Ability", "HP", "Atk", "Def", "SpA", "SpD", "Spe", "BST", "Sprite"]

    ws.append(header)

    for cell in ws["1:1"]:
        cell.font = Font(bold=True, underline="single")

    # Use a generator to create rows
    rows = (extract_pokemon_info(row) for row in pokemon_rows if
            not pokemon_list or row.find('td', class_='speciesNameWrapper').text.lower() in pokemon_list)
    for row in rows:
        ws.append(row)

    # Adjust column widths and alignment
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 21.5
    ws.column_dimensions['D'].width = 23.5
    ws.column_dimensions['E'].width = 23.5
    ws.column_dimensions['F'].width = 23.5
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12

    alignment = Alignment(horizontal='center')
    for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        for cell in ws[col]:
            cell.alignment = alignment

    ws.column_dimensions['B'].alignment = Alignment(horizontal='left')

    # Save the workbook
    xlsx_file_path = 'pokemon_data.xlsx'
    wb.save(xlsx_file_path)
    print("Saved Pokémon data to 'pokemon_data.xlsx'.")


def get_html_content():
    # Check if the HTML content file already exists
    if os.path.exists(html_file_path):
        print("HTML content file already exists.")
        with open(html_file_path, 'r', encoding='utf-8') as file:
            html_content = file.read()

        # Parse the HTML content
        soup = BeautifulSoup(html_content, 'lxml')

        # Check if the HTML content contains all the Pokémon in the list
        pokemon_names_in_html = [row.find('td', class_='speciesNameWrapper').text.lower() for row in
                                 soup.find_all('tr', class_='speciesRow')]
        if all(pokemon in pokemon_names_in_html for pokemon in pokemon_list):
            print("All Pokémon in the list are in the HTML content. Skipping to scraping part...")
            return html_content
        else:
            print("Not all Pokémon in the list are in the HTML content. Running the script to get the HTML content...")

    print("Starting the script...")
    # Path to your ChromeDriver executable
    chrome_driver_path = 'C:\\scripts\\chromedriver-win64\\chromedriver.exe'
    # Path to your Chrome executable
    chrome_binary_path = 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe'  # Change this to your Chrome installation path

    # Create a Service object
    service = Service(chrome_driver_path)

    # Set the Chrome binary location
    options = Options()
    options.binary_location = chrome_binary_path

    # Create a new instance of the Chrome driver with the specified options
    driver = webdriver.Chrome(service=service, options=options)

    # Go to the website
    driver.get('https://dex.radicalred.net/')
    print("Navigated to the website.")

    # Wait for the page to load completely
    time.sleep(2)

    # Define the file path
    file_path = 'C:\\Users\\Bhupen\\Desktop\\radical red\\Radical Red v4.1.sav'

    # Wait for the button to be clickable
    upload_button = WebDriverWait(driver, 4).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Upload save file (v4.1)')]"))
    )
    print("Found the upload button.")

    # Click the button to open the file dialog
    upload_button.click()
    print("Clicked the upload button.")

    # Wait for the file dialog to open
    time.sleep(2)  # Adjust this time according to your system's speed

    # Type the file path and press Enter using pyautogui
    pyautogui.write(file_path)
    pyautogui.press('enter')
    print("Uploaded the file.")

    # Wait for the page to load completely after file upload
    time.sleep(1)

    # Scroll down until the Pokémon with DexID 2001 is found
    print("Scrolling until DexID 2001 is found...")
    while True:
        # Scroll down
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Wait for the page to load
        time.sleep(0.1)  # Adjust this time to make the scrolling faster

        # Check if the Pokémon with DexID 2001 is present
        try:
            pokemon_2001 = driver.find_element(By.XPATH, "//td[@class='speciesDexIDWrapper' and text()='2001']")
            print("Found Pokémon with DexID 2001.")
            break  # If the Pokémon is found, break the loop
        except Exception:
            continue  # If the Pokémon is not found, continue scrolling

    # Wait for the table to load
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'speciesTable'))
    )
    print("Found the Pokémon table.")

    # Save the HTML content
    html_content = driver.page_source
    print("Saved the HTML content.")

    # Close the browser
    driver.quit()
    print("Closed the browser.")

    # Save the HTML content to a file
    with open(html_file_path, 'w', encoding='utf-8') as file:
        file.write(html_content)
    print(f"Saved HTML content to '{html_file_path}'.")

    return html_content


def ask_to_delete_html_file():
    # Ask the user if they want to delete the HTML content file
    delete_file = input(f"Do you want to delete the HTML content file '{html_file_path}'? (yes/no): ")
    if delete_file.lower() == 'yes':
        os.remove(html_file_path)
        print(f"Deleted the file '{html_file_path}'.")
    else:
        print(f"Kept the file '{html_file_path}'.")


if __name__ == "__main__":
    check_install_dependencies()

    # Get the HTML content
    html_content = get_html_content()

    # Scrape Pokémon data
    scrape_pokemon(html_content)

    # Ask to delete HTML file
    ask_to_delete_html_file()
